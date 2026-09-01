import openpyxl

file_path = ".shared/uploads/Accenture - Top Accounts - EMEA - FY26 (2).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

with open("analysis_summary.txt", "w") as out:
    out.write(f"Sheet names: {wb.sheetnames}\n")
    for name in wb.sheetnames:
        ws = wb[name]
        out.write(f"\n--- Sheet: {name} ---\n")
        count = 0
        for row in ws.iter_rows(values_only=True):
            if any(row):
                if count < 15:
                    out.write(f"Row {count+1}: {[str(c)[:40] if c is not None else None for c in row[:15]]}\n")
                count += 1
        out.write(f"Total non-empty rows: {count}\n")
